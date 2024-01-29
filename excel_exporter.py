import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Color, Alignment

def clear_worksheet_data(worksheet):
    # Iterar sobre las filas y columnas para borrar solo los valores, conservando los formatos
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.value = None

def dataframe_to_excel(worksheet, dataframe):

    clear_worksheet_data(worksheet)
    # Escribir los encabezados del DataFrame en la primera fila (fila 1)
    for col, columna in enumerate(dataframe.columns, 1):
        celda = worksheet.cell(row=1, column=col)
        celda.value = columna
        celda.font = Font(name='Ubuntu', bold=True,  color=Color(rgb="FFFFFF"))
        celda.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        celda.alignment = Alignment(horizontal="center", vertical="center")

    # Escribir los datos del DataFrame en las filas siguientes
    for fila, serie in enumerate(dataframe.itertuples(), 2):
        for col, valor in enumerate(serie[1:], 1):
            celda = worksheet.cell(row=fila, column=col)
            celda.value = valor

def export_to_excel(cedears_df, totals_df):
    wb = openpyxl.load_workbook(filename = 'archivo_formateado.xlsx')
    dataframe_to_excel(wb["Stock Performance"], cedears_df)
    dataframe_to_excel(wb["Totals"], totals_df)

    wb.save('my_performance.xlsx')